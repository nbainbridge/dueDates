# Header Section
__author__ = "Nicholas A. Bainbridge"
__version__ = "0.0.1 beta"

from openpyxl import load_workbook
from datetime import datetime
from cal_setup import get_calendar_service

xl_file = 'due_dates.xlsx'
dt_format = '%Y-%m-%d %H:%M:%S'
service = get_calendar_service()


def main():
    try:
        workbook = load_workbook(xl_file, data_only=True)  # Your Excel file
        worksheet = workbook.active  # gets first sheet
        row = 2
        while worksheet.cell(row=row, column=1).value is not None:
            if worksheet.cell(row=row, column=9).value != 'YES':
                try:
                    event_start = datetime.strptime(str(worksheet.cell(row=row, column=7).value), dt_format).isoformat()
                    event_end = datetime.strptime(str(worksheet.cell(row=row, column=8).value), dt_format).isoformat()
                    event = str(worksheet.cell(row=row, column=1).value)
                    event_desc = str(worksheet.cell(row=row, column=2).value)
                    worksheet.cell(row=row, column=9).value = "YES"
                    event_result = service.events().insert(
                        calendarId='8htsvne510csdegv4mlac881kc@group.calendar.google.com',
                        body={
                            "summary": event,
                            "description": event_desc,
                            "start": {"dateTime": event_start, "timeZone": 'Canada/Eastern'},
                            "end": {"dateTime": event_end, "timeZone": 'Canada/Eastern'},
                        }
                        ).execute()
                except Exception as error:
                    worksheet.cell(row=row, column=9).value = str(error)
            row += 1
        workbook.save(xl_file)
        workbook.close()
    except Exception as error:
        print("Error:")
        print(error)


if __name__ == '__main__':
    main()
